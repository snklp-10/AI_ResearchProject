import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import numbers

input_file = r"C:\Users\snklp\Downloads\ResearchProject\classification_GT\groundTruth_feline.json"
output_file = r"C:\Users\snklp\Downloads\CM.xlsx"

terms = [
    "pulmonary_nodules",
    "esophagitis",
    "pneumonia",
    "bronchitis",
    "interstitial",
    "diseased_lungs",
    "hypo_plastic_trachea",
    "cardiomegaly",
    "pleural_effusion",
    "perihilar_infiltrate",
    "rtm",
    "focal_caudodorsal_lung",
    "right_sided_cardiomegaly",
    "focal_perihilar",
    "left_sided_cardiomegaly",
    "bronchiectasis",
    "pulmonary_vessel_enlargement",
    "thoracic_lymphadenopathy",
    "pulmonary_hypoinflation",
    "pericardial_effusion",
    "Fe_Alveolar",
]

# Load JSON
with open(input_file, "r") as f:
    data = json.load(f)

# Initialize counts dictionary
counts = {term: {"fp": 0, "tp": 0, "fn": 0, "tn": 0} for term in terms}

# Count occurrences for each category
for entry in data:
    for col in ["tp", "fp", "fn", "tn"]:
        for term in entry.get(col, []):
            if term in terms:
                counts[term][col] += 1

# Handle output workbook and sheet naming
code_run_count = 1
while True:
    sheet_name = f"GT_CM_{code_run_count}"
    try:
        wb = load_workbook(output_file)
        if sheet_name in wb.sheetnames:
            code_run_count += 1
            continue
        break
    except FileNotFoundError:
        wb = None
        break

# Create workbook if needed
if wb is None:
    wb = Workbook()

ws = wb.create_sheet(sheet_name)

# Column headers
column_headers = [
    "Condition",
    "tp_Positive",
    "fn_Positive",
    "tn_Positive",
    "fp_Positive",
    "Sensitivity",
    "Specificity",
    "Check",
    "Positive Ground Truth",
    "Negative Ground Truth",
    "Ground Truth Check",
    "Radiologist Agreement Rate",
]
ws.append(column_headers)

# Fill sheet rows + formulas
row_start = 2
for i, (term, result) in enumerate(counts.items(), start=row_start):
    ws.cell(i, 1, term)
    ws.cell(i, 2, result["tp"])
    ws.cell(i, 3, result["fn"])
    ws.cell(i, 4, result["tn"])
    ws.cell(i, 5, result["fp"])

    ws.cell(i, 6, f"=IFERROR(B{i}/(B{i}+C{i}), 0)")  # Sensitivity
    ws.cell(i, 7, f"=IFERROR(D{i}/(D{i}+E{i}), 0)")  # Specificity
    ws.cell(i, 8, f"=SUM(B{i}:E{i})")  # Check
    ws.cell(i, 9, f"=SUM(B{i}:C{i})")  # Positive Ground Truth
    ws.cell(i, 10, f"=SUM(D{i}:E{i})")  # Negative Ground Truth
    ws.cell(i, 11, f"=SUM(I{i}:J{i})")  # Ground Truth Check
    ws.cell(i, 12, f"=IFERROR((B{i}+D{i})/H{i}, 0)")  # Radiologist Agreement Rate

    # Format percentages
    ws.cell(i, 6).number_format = numbers.FORMAT_PERCENTAGE_00
    ws.cell(i, 7).number_format = numbers.FORMAT_PERCENTAGE_00
    ws.cell(i, 12).number_format = numbers.FORMAT_PERCENTAGE_00

# Remove default sheet if blank
if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
    wb.remove(wb["Sheet"])

# Save
wb.save(output_file)
print(f"✅ Confusion matrix saved in sheet '{sheet_name}'")
