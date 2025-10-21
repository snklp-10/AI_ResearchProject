import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers

# import json

# from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r"C:\Users\snklp\Downloads\Research Student Assignments\Research Student Assignments\Input Data 1 - canine_thorax_scoring.xlsx"

output_file = r"C:\Users\snklp\Downloads\Research Student Assignments\Research Student Assignments\Example Confusion Matrix Output.xlsx"

thorax_terms = [
    "perihilar_infiltrate",
    "pneumonia",
    "bronchitis",
    "interstitial",
    "diseased_lungs",
    "hypo_plastic_trachea",
    "cardiomegaly",
    "pulmonary_nodules",
    "pleural_effusion",
    "rtm",
    "focal_caudodorsal_lung",
    "focal_perihilar",
    "pulmonary_hypoinflation",
    "right_sided_cardiomegaly",
    "pericardial_effusion",
    "bronchiectasis",
    "pulmonary_vessel_enlargement",
    "left_sided_cardiomegaly",
    "thoracic_lymphadenopathy",
    "esophagitis",
    "vhs_v2",
]
code_run_count = 1

while True:
    sheet_name = f"Confusion_Matrix_{code_run_count}"
    try:
        wb = load_workbook(output_file)
        if sheet_name in wb.sheetnames:
            code_run_count += 1
            continue
        break
    except FileNotFoundError:
        # If file doesn't exist, create a new workbook later
        wb = None
        break

try:
    # Counting values in the input file
    df = pd.read_excel(input_file, usecols=["fp", "tp", "fn", "tn"])
    df = df.fillna("")
    counts = {term: {"fp": 0, "tp": 0, "fn": 0, "tn": 0} for term in thorax_terms}

    for col in ["fp", "tp", "fn", "tn"]:
        for term in thorax_terms:
            counts[term][col] = df[col].astype(str).str.contains(term, case=False).sum()

    # for term, result in counts.items():
    #     print(f"{term}:{result}")

    # create new workbook if it doesn't exist
    if wb is None:
        from openpyxl import Workbook

        wb = Workbook()

    # Create new sheet for this run
    ws = wb.create_sheet(sheet_name)

    column_headers = [
        "Condition",
        "tp_Positive",
        "fn_Positive",
        "tn_Positive",
        "fp_Positive",
        "Sensitivity",
        "Specificity",
        "Check",
        "Positive Ground Truth",
        "Negative Ground Truth",
        "Ground Truth Check",
        "Radiologist Agreement Rate",
    ]
    ws.append(column_headers)

    row_start = 2
    for i, (term, result) in enumerate(counts.items(), start=row_start):
        ws.cell(i, 1, term)
        ws.cell(i, 2, result["tp"])
        ws.cell(i, 3, result["fn"])
        ws.cell(i, 4, result["tn"])
        ws.cell(i, 5, result["fp"])

        ws.cell(i, 6, f"=IFERROR(B{i}/(B{i}+C{i}), 0)")  # Sensitivity
        ws.cell(i, 7, f"=IFERROR(D{i}/(D{i}+E{i}), 0)")  # Specificity
        ws.cell(i, 8, f"=SUM(B{i}:E{i})")  # Check
        ws.cell(i, 9, f"=SUM(B{i}:C{i})")  # Positive Ground Truth
        ws.cell(i, 10, f"=SUM(D{i}:E{i})")  # Negative Ground Truth
        ws.cell(i, 11, f"=SUM(I{i}:J{i})")  # Grand  Truth Check
        ws.cell(i, 12, f"=IFERROR((B{i}+D{i})/H{i}, 0)")  # Radiologist Agreement Rate

        ws.cell(i, 6).number_format = numbers.FORMAT_PERCENTAGE_00
        ws.cell(i, 7).number_format = numbers.FORMAT_PERCENTAGE_00
        ws.cell(i, 12).number_format = numbers.FORMAT_PERCENTAGE_00

    # remove sheet if it has no data
    if (
        "Sheet" in wb.sheetnames
        and wb["Sheet"].max_row == 1
        and wb["Sheet"].max_column == 1
    ):
        wb.remove(wb["Sheet"])

    # save file
    wb.save(output_file)
    print(f"✅ Confusion matrix saved in sheet '{sheet_name}'")

except FileNotFoundError:
    print(f"Error: The file {input_file} was not found")
except Exception as e:
    print(f"An error occurred {e}")
